import logging
import csv
import re
from datetime import datetime

from typing import List
from sqlalchemy import select, exists, update, and_, func
from sqlalchemy.exc import SQLAlchemyError

from database.database import get_db_session
from database.models import Post, Channel, ChannelHistory, Blacklist

from constants.db_constants import DEFAULT_PATTERNS
from constants.logger import LOG_DB


logger = logging.getLogger(__name__)


async def initialize_blacklist():
    async with get_db_session() as session:
        try:
            for pattern, reason in DEFAULT_PATTERNS:
                result = await session.execute(
                    select(
                        exists().where(
                            and_(
                                Blacklist.pattern == pattern, Blacklist.reason == reason
                            )
                        )
                    )
                )
                blacklist_db = result.scalar()
                if not blacklist_db:
                    blacklist = Blacklist(pattern=pattern, reason=reason)
                    session.add(blacklist)
                    await session.commit()
                    await session.flush()
                    logger.info(LOG_DB["create_blacklist"])
            return True
        except SQLAlchemyError as e:
            logging.error(LOG_DB["db_err"].format(error=e))
            return False


async def get_unchecked_posts_count():
    async with get_db_session() as session:
        try:
            result = await session.execute(
                select(func.count()).select_from(Post).where(Post.is_processed == False)
            )
            count = result.scalar()
            return count or 0
        except SQLAlchemyError as e:
            logging.error(e)
            return 0  # Return 0 instead of False for consistency


async def   export_data_to_csv():
    try:
        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        async with get_db_session() as session:
            headers = [col.name for col in Post.__table__.columns]

            result = await session.execute(select(Post))
            posts = result.scalars().all()
        with open(filename, mode="w", newline="", encoding="utf-8-sig") as file:
            writer = csv.writer(file, delimiter=";", quoting=csv.QUOTE_MINIMAL)

            writer.writerow(headers)

            for post in posts:
                row = [getattr(post, col.name) for col in Post.__table__.columns]
                cleaned_row = [
                    (
                        str(item).replace(";", ",").strip()
                        if isinstance(item, str)
                        else item
                    )
                    for item in row
                ]
                writer.writerow(cleaned_row)
            logger.info(LOG_DB["export_csv"])
        return filename
    except Exception as e:
        logging.error(LOG_DB["db_err"].format(error=e))
        return False


async def export_data_to_excel():
    try:
        from openpyxl import Workbook
        from datetime import datetime
        
        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        async with get_db_session() as session:
            headers = [col.name for col in Post.__table__.columns]
            
            result = await session.execute(select(Post))
            posts = result.scalars().all()
        
        # Создаем новую книгу Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Posts"
        
        # Записываем заголовки
        for col_idx, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_idx).value = header
        
        # Записываем данные
        for row_idx, post in enumerate(posts, 2):
            for col_idx, col in enumerate(Post.__table__.columns, 1):
                value = getattr(post, col.name)
                # Преобразуем значение в строку, если это необходимо
                if isinstance(value, str):
                    value = value.strip()
                sheet.cell(row=row_idx, column=col_idx).value = value
        
        # Сохраняем файл
        workbook.save(filename)
        logger.info(LOG_DB["export_csv"])
        return filename
    
    except Exception as e:
        logging.error(LOG_DB["db_err"].format(error=e))
        return False

async def save_post(
    check_date, post_date, channel_link, post_link, post_text, user_requested=0
):
    async with get_db_session() as session:
        try:
            post = Post(
                check_date=check_date,
                post_date=post_date,
                channel_link=channel_link,
                post_link=post_link,
                post_text=post_text,
                user_requested=user_requested,
            )

            result = await session.execute(
                select(
                    exists().where(
                        and_(Post.post_text == post_text, Post.post_link == post_link)
                    )
                )
            )
            post_db = result.scalar()
            if not post_db:
                session.add(post)
                await session.commit()
                await session.refresh(post)
                return True
            return False
        except SQLAlchemyError as e:
            logger.error(LOG_DB["db_err"].format(error=e))
            return False


async def add_channel(channel_link, source="parser"):
    # Normalize channel link
    if "@" in channel_link:
        channel_link = channel_link.strip()
    elif "t.me/" in channel_link:
        channel_link = "@" + channel_link.split("t.me/")[-1].strip()
    else:
        channel_link = "@" + channel_link.split("/")[-1].strip()
    
    async with get_db_session() as session:
        try:
            result = await session.execute(
                select(Channel).where(Channel.channel_link == channel_link)
            )
            channel = result.scalar_one_or_none()
            if not channel:
                new_channel = Channel(channel_link=channel_link, source=source)
                session.add(new_channel)
                await session.commit()
                await session.refresh(new_channel)
                return True
            channel.is_active = True
            await session.commit()
            await session.refresh(channel)
            return True
        except SQLAlchemyError as e:
            logger.error(LOG_DB["db_err"].format(error=e))
            return False


async def is_blacklisted(value: str, check_pattern: bool = False):
    async with get_db_session() as session:
        try:
            if check_pattern:
                result = await session.execute(select(Blacklist.pattern))
                patterns = result.scalars().all()
                for pattern in patterns:
                    if re.search(pattern, value):
                        return True
                return False
            else:
                result = await session.execute(
                    select(Blacklist).where(Blacklist.pattern == value)
                )
                pattern_db = result.scalar_one_or_none()
                return pattern_db is not None
        except SQLAlchemyError as e:
            logger.error(LOG_DB["db_err"].format(error=e))
            return False


async def add_to_blacklist(pattern: str, reason: str = "") -> bool:
    async with get_db_session() as session:
        try:
            blacklist = Blacklist(pattern=pattern, reason=reason)
            session.add(blacklist)
            await session.commit()
            await session.refresh(blacklist)
            return True
        except SQLAlchemyError as e:
            logger.error(LOG_DB["db_err"].format(error=e))
            return False


async def save_new_channels(channels: List[str], source: str = "auto_find") -> int:
    saved_count = 0
    async with get_db_session() as session:
        try:
            # Fix the syntax error in the where clause
            existing = await session.execute(
                select(Channel.channel_link).where(Channel.channel_link.in_(channels))
            )
            existing_links = set(existing.scalars().all())
            
            for channel in channels:
                if channel not in existing_links:
                    # Fix the URL formatting (remove space after t.me/)
                    link = channel if channel.startswith('@') else f"@{channel}"
                    new_channel = Channel(
                        channel_link=link, added_date=datetime.now(), source=source
                    )
                    session.add(new_channel)
                    await session.commit()
                    await session.refresh(new_channel)
                    saved_count += 1
            return saved_count
        except SQLAlchemyError as e:
            logger.error(LOG_DB["db_err"].format(error=e))
            return 0  # Return 0 instead of False for consistency


async def mark_post_as_checked(post_id, is_recipe):
    async with get_db_session() as session:
        try:
            stmt = update(Post).where(Post.id == post_id).values(is_processed=True, is_recipe=is_recipe)
            await session.execute(stmt)
            await session.commit()
            return True
        except SQLAlchemyError as e:
            logger.error(LOG_DB["db_err"].format(error=e))
            return False


async def get_unchecked_posts(limit=None):
    async with get_db_session() as session:
        try:
            query = select(Post.id, Post.post_text).where(Post.is_processed == False)
            if limit:
                query = query.limit(limit)
            result = await session.execute(query)
            posts = result.all()
            return posts
        except SQLAlchemyError as e:
            logger.error(LOG_DB["db_err"].format(error=e))
            return []  # Return empty list instead of False for consistency


async def get_active_channels():
    async with get_db_session() as session:
        try:
            result = await session.execute(
                select(Channel.channel_link).where(Channel.is_active == True)
            )
            channels = result.scalars().all()
            return channels
        except SQLAlchemyError as e:
            logger.error(LOG_DB["db_err"].format(error=e))
            return []


async def deactivate_channel(channel_link: str, error_message: str = ""):
    async with get_db_session() as session:
        try:
            stmt = update(Channel).where(Channel.channel_link == channel_link).values(is_active=False)
            await session.execute(stmt)
            
            history = ChannelHistory(
                channel_link=channel_link,
                status="inactive",
                error_message=error_message[:500],
            )
            session.add(history)
            await session.commit()
            await session.refresh(history)
            return True
        except SQLAlchemyError as e:
            logger.error(LOG_DB["db_err"].format(error=e))
            return False


async def get_stats():
    async with get_db_session() as session:
        try:
            result = await session.execute(select(func.count()).select_from(Post))
            all_posts = result.scalar() or 0

            result = await session.execute(
                select(func.count()).select_from(Post).where(Post.is_recipe == True)
            )
            recipe_posts = result.scalar() or 0

            result = await session.execute(
                select(func.count()).select_from(Post).where(Post.is_processed == False)
            )
            unck_posts = result.scalar() or 0

            final_dict = {
                "total_posts": all_posts,
                "recipes": recipe_posts,
                "unchecked": unck_posts,
            }
            return final_dict
        except SQLAlchemyError as e:
            logger.error(LOG_DB["db_err"].format(error=e))
            return {"total_posts": 0, "recipes": 0, "unchecked": 0}  # Return default dict instead of False


async def get_posts_for_search():
    async with get_db_session() as session:
        try:
            result = await session.execute(
                select(Post.post_text).where(Post.is_processed == False)
            )
            posts = result.all()  # Use .all() instead of .scalars().all() to match expected return format
            return posts
        except SQLAlchemyError as e:
            logger.error(LOG_DB["db_err"].format(error=e))
            return []


async def get_channel_links():
    async with get_db_session() as session:
        try:
            result = await session.execute(select(Channel.channel_link))
            links = result.scalars().all()
            return {link.lower() for link in links}
        except SQLAlchemyError as e:
            logger.error(LOG_DB["db_err"].format(error=e))
            return set()


async def get_blacklist_pat_reason():
    async with get_db_session() as session:
        try:
            result = await session.execute(
                select(Blacklist.pattern, Blacklist.reason)
                .order_by(Blacklist.added_date.desc())
                .limit(50)
            )
            blacklist = result.all()
            return blacklist
        except SQLAlchemyError as e:
            logger.error(LOG_DB["db_err"].format(error=e))
            return []


async def insert_repl_chan_history(channel_link: str):
    async with get_db_session() as session:
        try:
            result = await session.execute(
                select(ChannelHistory).where(
                    ChannelHistory.channel_link == channel_link
                )
            )
            channel = result.scalar_one_or_none()
            if not channel:
                new_channel = ChannelHistory(channel_link=channel_link, status="active")
                session.add(new_channel)
                await session.commit()
                await session.refresh(new_channel)
                return True
            else:
                stmt = update(ChannelHistory).where(
                    ChannelHistory.channel_link == channel_link
                ).values(status="active")
                await session.execute(stmt)
                await session.commit()
                return True

        except SQLAlchemyError as e:
            logger.error(LOG_DB["db_err"].format(error=e))
            return False
